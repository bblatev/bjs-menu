"""Scheduled Reports Service.

Allows scheduling automatic report generation and email delivery.
Supports daily, weekly, and monthly reports in PDF and Excel formats.
"""

import logging
import asyncio
from datetime import datetime, date, time, timedelta
from typing import Optional, List, Dict, Any, Callable
from dataclasses import dataclass, field
from enum import Enum
from io import BytesIO
import json

logger = logging.getLogger(__name__)


class ReportFrequency(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    CUSTOM = "custom"  # Custom cron-like schedule


class ReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"


class ReportType(str, Enum):
    DAILY_SALES = "daily_sales"
    WEEKLY_SALES = "weekly_sales"
    MONTHLY_SALES = "monthly_sales"
    INVENTORY_STATUS = "inventory_status"
    LABOR_SUMMARY = "labor_summary"
    MENU_PERFORMANCE = "menu_performance"
    CUSTOMER_INSIGHTS = "customer_insights"
    FINANCIAL_SUMMARY = "financial_summary"
    CUSTOM = "custom"


@dataclass
class ReportSchedule:
    """A scheduled report configuration."""
    schedule_id: str
    name: str
    report_type: ReportType
    frequency: ReportFrequency
    format: ReportFormat
    recipients: List[str]  # Email addresses
    venue_id: Optional[int] = None

    # Schedule details
    time_of_day: time = field(default_factory=lambda: time(6, 0))  # 6 AM default
    day_of_week: Optional[int] = None  # 0=Monday, 6=Sunday (for weekly)
    day_of_month: Optional[int] = None  # 1-28 (for monthly)

    # Report parameters
    parameters: Dict[str, Any] = field(default_factory=dict)

    # Status
    is_active: bool = True
    last_run: Optional[datetime] = None
    next_run: Optional[datetime] = None
    last_status: str = "pending"
    last_error: Optional[str] = None

    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime = field(default_factory=datetime.utcnow)


@dataclass
class ReportRun:
    """A record of a report execution."""
    run_id: str
    schedule_id: str
    report_type: ReportType
    started_at: datetime
    completed_at: Optional[datetime] = None
    status: str = "running"  # running, success, failed
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    recipients_notified: int = 0
    error_message: Optional[str] = None


class ScheduledReportsService:
    """Service for managing scheduled reports."""

    def __init__(
        self,
        notification_service=None,
        export_service=None,
    ):
        self.notification_service = notification_service
        self.export_service = export_service

        # In-memory storage (use database in production)
        self._schedules: Dict[str, ReportSchedule] = {}
        self._runs: Dict[str, ReportRun] = {}

        # Report generators
        self._generators: Dict[ReportType, Callable] = {}

    # =========================================================================
    # Schedule Management
    # =========================================================================

    def create_schedule(
        self,
        name: str,
        report_type: ReportType,
        frequency: ReportFrequency,
        format: ReportFormat,
        recipients: List[str],
        time_of_day: time = time(6, 0),
        day_of_week: Optional[int] = None,
        day_of_month: Optional[int] = None,
        venue_id: Optional[int] = None,
        parameters: Optional[Dict[str, Any]] = None,
    ) -> ReportSchedule:
        """Create a new report schedule."""
        schedule_id = f"RS-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

        schedule = ReportSchedule(
            schedule_id=schedule_id,
            name=name,
            report_type=report_type,
            frequency=frequency,
            format=format,
            recipients=recipients,
            venue_id=venue_id,
            time_of_day=time_of_day,
            day_of_week=day_of_week,
            day_of_month=day_of_month,
            parameters=parameters or {},
        )

        # Calculate next run time
        schedule.next_run = self._calculate_next_run(schedule)

        self._schedules[schedule_id] = schedule
        logger.info(f"Created report schedule {schedule_id}: {name}")

        return schedule

    def update_schedule(
        self,
        schedule_id: str,
        **updates,
    ) -> Optional[ReportSchedule]:
        """Update a report schedule."""
        if schedule_id not in self._schedules:
            return None

        schedule = self._schedules[schedule_id]

        for key, value in updates.items():
            if hasattr(schedule, key) and value is not None:
                setattr(schedule, key, value)

        schedule.updated_at = datetime.utcnow()
        schedule.next_run = self._calculate_next_run(schedule)

        logger.info(f"Updated report schedule {schedule_id}")
        return schedule

    def delete_schedule(self, schedule_id: str) -> bool:
        """Delete a report schedule."""
        if schedule_id not in self._schedules:
            return False

        del self._schedules[schedule_id]
        logger.info(f"Deleted report schedule {schedule_id}")
        return True

    def get_schedule(self, schedule_id: str) -> Optional[ReportSchedule]:
        """Get a report schedule."""
        return self._schedules.get(schedule_id)

    def list_schedules(
        self,
        venue_id: Optional[int] = None,
        is_active: Optional[bool] = None,
    ) -> List[ReportSchedule]:
        """List all report schedules."""
        schedules = list(self._schedules.values())

        if venue_id is not None:
            schedules = [s for s in schedules if s.venue_id == venue_id]

        if is_active is not None:
            schedules = [s for s in schedules if s.is_active == is_active]

        return sorted(schedules, key=lambda s: s.next_run or datetime.max)

    def toggle_schedule(self, schedule_id: str, is_active: bool) -> Optional[ReportSchedule]:
        """Enable or disable a schedule."""
        schedule = self._schedules.get(schedule_id)
        if not schedule:
            return None

        schedule.is_active = is_active
        schedule.updated_at = datetime.utcnow()

        if is_active:
            schedule.next_run = self._calculate_next_run(schedule)

        return schedule

    # =========================================================================
    # Next Run Calculation
    # =========================================================================

    def _calculate_next_run(self, schedule: ReportSchedule) -> datetime:
        """Calculate the next run time for a schedule."""
        now = datetime.utcnow()
        run_time = datetime.combine(now.date(), schedule.time_of_day)

        if schedule.frequency == ReportFrequency.DAILY:
            if run_time <= now:
                run_time += timedelta(days=1)

        elif schedule.frequency == ReportFrequency.WEEKLY:
            target_day = schedule.day_of_week or 0  # Monday default
            days_ahead = target_day - now.weekday()
            if days_ahead < 0 or (days_ahead == 0 and run_time <= now):
                days_ahead += 7
            run_time = datetime.combine(
                now.date() + timedelta(days=days_ahead),
                schedule.time_of_day
            )

        elif schedule.frequency == ReportFrequency.MONTHLY:
            target_day = schedule.day_of_month or 1
            # Find next occurrence of target day
            if now.day > target_day or (now.day == target_day and run_time <= now):
                # Move to next month
                if now.month == 12:
                    next_month = datetime(now.year + 1, 1, target_day)
                else:
                    next_month = datetime(now.year, now.month + 1, target_day)
                run_time = datetime.combine(next_month.date(), schedule.time_of_day)
            else:
                run_time = datetime.combine(
                    datetime(now.year, now.month, target_day).date(),
                    schedule.time_of_day
                )

        return run_time

    # =========================================================================
    # Report Generation
    # =========================================================================

    def register_generator(
        self,
        report_type: ReportType,
        generator: Callable,
    ):
        """Register a report generator function."""
        self._generators[report_type] = generator
        logger.info(f"Registered generator for {report_type}")

    async def run_report(
        self,
        schedule_id: str,
        force: bool = False,
    ) -> Optional[ReportRun]:
        """Run a scheduled report."""
        schedule = self._schedules.get(schedule_id)
        if not schedule:
            logger.error(f"Schedule {schedule_id} not found")
            return None

        if not schedule.is_active and not force:
            logger.warning(f"Schedule {schedule_id} is not active")
            return None

        # Create run record
        run_id = f"RR-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        run = ReportRun(
            run_id=run_id,
            schedule_id=schedule_id,
            report_type=schedule.report_type,
            started_at=datetime.utcnow(),
        )
        self._runs[run_id] = run

        try:
            # Generate report data
            generator = self._generators.get(schedule.report_type)
            if not generator:
                raise ValueError(f"No generator for {schedule.report_type}")

            report_data = await generator(schedule.parameters)

            # Export to file
            file_content, file_name = await self._export_report(
                report_data,
                schedule.report_type,
                schedule.format,
            )

            run.file_path = file_name
            run.file_size = len(file_content) if file_content else 0

            # Send to recipients
            if schedule.recipients and self.notification_service:
                sent_count = await self._send_report_email(
                    schedule.recipients,
                    schedule.name,
                    file_content,
                    file_name,
                    schedule.format,
                )
                run.recipients_notified = sent_count

            # Update run record
            run.completed_at = datetime.utcnow()
            run.status = "success"

            # Update schedule
            schedule.last_run = run.started_at
            schedule.last_status = "success"
            schedule.last_error = None
            schedule.next_run = self._calculate_next_run(schedule)

            logger.info(f"Successfully ran report {schedule_id}")

        except Exception as e:
            run.completed_at = datetime.utcnow()
            run.status = "failed"
            run.error_message = str(e)

            schedule.last_run = run.started_at
            schedule.last_status = "failed"
            schedule.last_error = str(e)

            logger.error(f"Failed to run report {schedule_id}: {e}")

        return run

    async def _export_report(
        self,
        data: Dict[str, Any],
        report_type: ReportType,
        format: ReportFormat,
    ) -> tuple:
        """Export report data to file."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_name = f"{report_type.value}_{timestamp}"

        if format == ReportFormat.PDF:
            # Use reportlab for PDF generation
            content = await self._generate_pdf(data, report_type)
            file_name += ".pdf"

        elif format == ReportFormat.EXCEL:
            # Use openpyxl for Excel
            content = await self._generate_excel(data, report_type)
            file_name += ".xlsx"

        elif format == ReportFormat.CSV:
            content = await self._generate_csv(data, report_type)
            file_name += ".csv"

        else:
            content = json.dumps(data, indent=2, default=str).encode()
            file_name += ".json"

        return content, file_name

    async def _generate_pdf(
        self,
        data: Dict[str, Any],
        report_type: ReportType,
    ) -> bytes:
        """Generate PDF report."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title = Paragraph(
                f"<b>{report_type.value.replace('_', ' ').title()} Report</b>",
                styles['Heading1']
            )
            story.append(title)
            story.append(Spacer(1, 12))

            # Generated timestamp
            story.append(Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                styles['Normal']
            ))
            story.append(Spacer(1, 20))

            # Data tables
            for section, section_data in data.items():
                story.append(Paragraph(f"<b>{section}</b>", styles['Heading2']))
                story.append(Spacer(1, 6))

                if isinstance(section_data, list) and section_data:
                    # Create table from list of dicts
                    if isinstance(section_data[0], dict):
                        headers = list(section_data[0].keys())
                        table_data = [headers]
                        for row in section_data[:50]:  # Limit rows
                            table_data.append([str(row.get(h, '')) for h in headers])

                        t = Table(table_data)
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        story.append(t)
                    else:
                        for item in section_data[:20]:
                            story.append(Paragraph(f"• {item}", styles['Normal']))

                elif isinstance(section_data, dict):
                    for key, value in section_data.items():
                        story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))

                else:
                    story.append(Paragraph(str(section_data), styles['Normal']))

                story.append(Spacer(1, 15))

            doc.build(story)
            return buffer.getvalue()

        except ImportError:
            logger.warning("reportlab not installed, returning JSON instead")
            return json.dumps(data, indent=2, default=str).encode()

    async def _generate_excel(
        self,
        data: Dict[str, Any],
        report_type: ReportType,
    ) -> bytes:
        """Generate Excel report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = report_type.value[:31]  # Max 31 chars

            row = 1

            # Title
            ws.cell(row=row, column=1, value=f"{report_type.value.replace('_', ' ').title()} Report")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 2

            # Generated timestamp
            ws.cell(row=row, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}")
            row += 2

            # Data sections
            for section, section_data in data.items():
                ws.cell(row=row, column=1, value=section)
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

                if isinstance(section_data, list) and section_data:
                    if isinstance(section_data[0], dict):
                        headers = list(section_data[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=header)
                            ws.cell(row=row, column=col).fill = PatternFill(
                                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                            )
                        row += 1

                        for item in section_data:
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=item.get(header, ''))
                            row += 1

                elif isinstance(section_data, dict):
                    for key, value in section_data.items():
                        ws.cell(row=row, column=1, value=key)
                        ws.cell(row=row, column=2, value=str(value))
                        row += 1

                row += 1

            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            logger.warning("openpyxl not installed, returning JSON instead")
            return json.dumps(data, indent=2, default=str).encode()

    async def _generate_csv(
        self,
        data: Dict[str, Any],
        report_type: ReportType,
    ) -> bytes:
        """Generate CSV report."""
        import csv

        buffer = BytesIO()
        writer = csv.writer(buffer)

        for section, section_data in data.items():
            writer.writerow([f"=== {section} ==="])

            if isinstance(section_data, list) and section_data:
                if isinstance(section_data[0], dict):
                    headers = list(section_data[0].keys())
                    writer.writerow(headers)
                    for item in section_data:
                        writer.writerow([item.get(h, '') for h in headers])
                else:
                    for item in section_data:
                        writer.writerow([item])

            elif isinstance(section_data, dict):
                for key, value in section_data.items():
                    writer.writerow([key, value])

            writer.writerow([])

        return buffer.getvalue()

    async def _send_report_email(
        self,
        recipients: List[str],
        report_name: str,
        file_content: bytes,
        file_name: str,
        format: ReportFormat,
    ) -> int:
        """Send report via email."""
        if not self.notification_service:
            logger.warning("No notification service configured")
            return 0

        # This would use the notification service to send emails with attachments
        # For now, just log
        logger.info(f"Would send {report_name} to {len(recipients)} recipients")
        return len(recipients)

    # =========================================================================
    # Run History
    # =========================================================================

    def get_run_history(
        self,
        schedule_id: Optional[str] = None,
        limit: int = 50,
    ) -> List[ReportRun]:
        """Get report run history."""
        runs = list(self._runs.values())

        if schedule_id:
            runs = [r for r in runs if r.schedule_id == schedule_id]

        return sorted(runs, key=lambda r: r.started_at, reverse=True)[:limit]

    # =========================================================================
    # Scheduler (Background Task)
    # =========================================================================

    async def check_and_run_due_reports(self):
        """Check for due reports and run them."""
        now = datetime.utcnow()

        for schedule in self._schedules.values():
            if not schedule.is_active:
                continue

            if schedule.next_run and schedule.next_run <= now:
                logger.info(f"Running due report: {schedule.schedule_id}")
                await self.run_report(schedule.schedule_id)


# Singleton instance
_reports_service: Optional[ScheduledReportsService] = None


def get_scheduled_reports_service() -> ScheduledReportsService:
    """Get the scheduled reports service singleton."""
    global _reports_service
    if _reports_service is None:
        _reports_service = ScheduledReportsService()

        # Register default generators
        async def daily_sales_generator(params):
            # Mock data - in production, query database
            return {
                "summary": {
                    "total_revenue": 5432.50,
                    "total_orders": 127,
                    "average_ticket": 42.78,
                },
                "hourly_breakdown": [
                    {"hour": h, "revenue": 200 + h * 50, "orders": 5 + h}
                    for h in range(11, 23)
                ],
            }

        _reports_service.register_generator(ReportType.DAILY_SALES, daily_sales_generator)

    return _reports_service
