"""Order service: Export functions for purchase orders."""

import io
from datetime import datetime
from typing import Dict

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.models.order import PurchaseOrder
from app.models.product import Product
from app.models.supplier import Supplier


def _prefetch_order_data(order: PurchaseOrder, db: Session) -> tuple[Supplier | None, Dict[int, Product]]:
    """Prefetch supplier and products to avoid N+1 queries."""
    # Fetch supplier
    supplier = db.query(Supplier).filter(Supplier.id == order.supplier_id).first()

    # Batch fetch all products for order lines
    product_ids = [line.product_id for line in order.lines if line.product_id]
    products = {}
    if product_ids:
        product_list = db.query(Product).filter(Product.id.in_(product_ids)).all()
        products = {p.id: p for p in product_list}

    return supplier, products


def generate_whatsapp_text(order: PurchaseOrder, db: Session) -> str:
    """Generate WhatsApp-ready text for a purchase order."""
    supplier, products = _prefetch_order_data(order, db)

    lines = []
    lines.append(f"*PURCHASE ORDER #{order.id}*")
    lines.append(f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}")
    if supplier:
        lines.append(f"Supplier: {supplier.name}")
    lines.append("")
    lines.append("*Items:*")

    total_items = 0
    for line in order.lines:
        product = products.get(line.product_id)
        if product:
            lines.append(f"• {product.name}: {line.qty} {product.unit}")
            total_items += 1

    lines.append("")
    lines.append(f"Total items: {total_items}")

    if order.notes:
        lines.append("")
        lines.append(f"Notes: {order.notes}")

    return "\n".join(lines)


def generate_pdf(order: PurchaseOrder, db: Session) -> bytes:
    """Generate PDF for a purchase order."""
    supplier, products = _prefetch_order_data(order, db)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=20,
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Purchase Order #{order.id}", title_style))

    # Supplier info
    if supplier:
        elements.append(Paragraph(f"<b>Supplier:</b> {supplier.name}", styles["Normal"]))
        if supplier.contact_phone:
            elements.append(Paragraph(f"<b>Phone:</b> {supplier.contact_phone}", styles["Normal"]))
        if supplier.contact_email:
            elements.append(Paragraph(f"<b>Email:</b> {supplier.contact_email}", styles["Normal"]))

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"<b>Date:</b> {order.created_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Status:</b> {order.status.value}", styles["Normal"]))

    elements.append(Spacer(1, 1 * cm))

    # Items table
    table_data = [["#", "Product", "Barcode", "Qty", "Unit", "Unit Cost", "Total"]]

    total_cost = 0
    for idx, line in enumerate(order.lines, 1):
        product = products.get(line.product_id)
        if product:
            unit_cost = line.unit_cost or product.cost_price or 0
            line_total = float(line.qty) * float(unit_cost)
            total_cost += line_total
            table_data.append([
                str(idx),
                product.name[:30],  # Truncate long names
                product.barcode or "",
                str(line.qty),
                product.unit,
                f"{unit_cost:.2f}" if unit_cost else "-",
                f"{line_total:.2f}" if unit_cost else "-",
            ])

    # Add total row
    table_data.append(["", "", "", "", "", "Total:", f"{total_cost:.2f}"])

    table = Table(table_data, colWidths=[1 * cm, 6 * cm, 3 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -2), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))

    elements.append(table)

    # Notes
    if order.notes:
        elements.append(Spacer(1, 1 * cm))
        elements.append(Paragraph(f"<b>Notes:</b> {order.notes}", styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()


def generate_xlsx(order: PurchaseOrder, db: Session) -> bytes:
    """Generate Excel file for a purchase order."""
    supplier, products = _prefetch_order_data(order, db)

    wb = Workbook()
    ws = wb.active
    ws.title = f"PO-{order.id}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header info
    ws["A1"] = f"Purchase Order #{order.id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    ws["A3"] = "Supplier:"
    ws["B3"] = supplier.name if supplier else ""
    ws["A4"] = "Date:"
    ws["B4"] = order.created_at.strftime("%Y-%m-%d %H:%M")
    ws["A5"] = "Status:"
    ws["B5"] = order.status.value

    # Items header
    headers = ["#", "Product", "Barcode", "Qty", "Unit", "Unit Cost", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Items
    total_cost = 0
    for idx, line in enumerate(order.lines, 1):
        row = 7 + idx
        product = products.get(line.product_id)
        if product:
            unit_cost = float(line.unit_cost or product.cost_price or 0)
            line_total = float(line.qty) * unit_cost
            total_cost += line_total

            data = [idx, product.name, product.barcode or "", float(line.qty), product.unit, unit_cost, line_total]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

    # Total row
    total_row = 8 + len(order.lines)
    ws.cell(row=total_row, column=6, value="Total:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_cost).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
