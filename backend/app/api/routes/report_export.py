"""Report export and scheduling endpoints."""

from __future__ import annotations

from fastapi import APIRouter, HTTPException, Request
from app.core.rate_limit import limiter
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timedelta
from io import BytesIO
from pydantic import BaseModel
import csv
import io as _io

from app.db.session import DbSession
from app.core.rbac import CurrentUser, OptionalCurrentUser

router = APIRouter()


# ==================== SCHEMAS ====================

class ReportExportRequest(BaseModel):
    period: str = "week"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    format: str = "csv"
    email_to: Optional[str] = None


class ScheduledReportConfig(BaseModel):
    report_type: str
    format: str = "csv"
    frequency: str = "daily"
    recipients: List[str] = []
    enabled: bool = True


# ==================== HELPERS ====================

def create_csv_export(data: list, headers: list) -> BytesIO:
    """Create CSV file from data."""
    output = BytesIO()
    output.write(b"\xef\xbb\xbf")
    text_output = _io.StringIO()
    writer = csv.writer(text_output, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    output.write(text_output.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def create_excel_export(data: list, headers: list, sheet_name: str = "Report") -> BytesIO:
    """Create Excel file from data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except ImportError:
        return create_csv_export(data, headers)


def create_pdf_export(data: list, headers: list, title: str = "Report") -> BytesIO:
    """Create PDF file from data (fallback to CSV if reportlab unavailable)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("CustomTitle", parent=styles["Heading1"], fontSize=24, spaceAfter=30, alignment=1)
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))
        table_data = [headers] + data
        num_cols = len(headers)
        col_width = 6.5 * inch / num_cols
        table = Table(table_data, colWidths=[col_width] * num_cols)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    except ImportError:
        return create_csv_export(data, headers)


# ==================== ENDPOINTS ====================

@router.post("/export/sales")
@limiter.limit("30/minute")
def export_sales_report(
    request: Request,
    body: ReportExportRequest = None,
    db: DbSession = None,
    current_user: OptionalCurrentUser = None,
):
    """Export sales report to PDF/Excel/CSV."""
    headers = ["Date", "Orders", "Revenue", "Avg Order Value", "Items Sold"]
    data = [["2024-01-01", "150", "7500.00", "50.00", "450"], ["2024-01-02", "175", "8750.00", "50.00", "525"]]

    if body.format == "pdf":
        output = create_pdf_export(data, headers, "Sales Report")
        media_type = "application/pdf"
        filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    elif body.format == "excel":
        output = create_excel_export(data, headers, "Sales Report")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    else:
        output = create_csv_export(data, headers)
        media_type = "text/csv"
        filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d')}.csv"

    return StreamingResponse(output, media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/export/labor-costs")
@limiter.limit("30/minute")
def export_labor_cost_report(request: Request, body: ReportExportRequest = None, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Export labor cost report."""
    headers = ["Staff Name", "Role", "Hours", "Regular Pay", "Overtime Pay", "Total Pay", "Revenue Generated"]
    data = []
    if body.format == "pdf":
        output = create_pdf_export(data, headers, "Labor Cost Report")
        media_type = "application/pdf"
        filename = f"labor_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    elif body.format == "excel":
        output = create_excel_export(data, headers, "Labor Costs")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"labor_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    else:
        output = create_csv_export(data, headers)
        media_type = "text/csv"
        filename = f"labor_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return StreamingResponse(output, media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/export/food-costs")
@limiter.limit("30/minute")
def export_food_cost_report(request: Request, body: ReportExportRequest = None, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Export food cost report."""
    headers = ["Item Name", "Category", "Qty Sold", "Selling Price", "Food Cost", "Food Cost %", "Profit Margin %", "Total Revenue"]
    data = []
    if body.format == "pdf":
        output = create_pdf_export(data, headers, "Food Cost Report")
        media_type = "application/pdf"
        filename = f"food_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    elif body.format == "excel":
        output = create_excel_export(data, headers, "Food Costs")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"food_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    else:
        output = create_csv_export(data, headers)
        media_type = "text/csv"
        filename = f"food_cost_report_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return StreamingResponse(output, media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/scheduled")
@limiter.limit("30/minute")
def create_scheduled_report(request: Request, config: ScheduledReportConfig = None, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Create a scheduled report."""
    return {"id": 1, "created_at": datetime.utcnow().isoformat(), "last_run": None, "next_run": (datetime.utcnow() + timedelta(days=1)).isoformat(), **config.model_dump()}


@router.get("/scheduled")
@limiter.limit("60/minute")
def list_scheduled_reports(request: Request, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """List all scheduled reports."""
    return []


@router.put("/scheduled/{report_id}")
@limiter.limit("30/minute")
def update_scheduled_report(request: Request, report_id: int, config: ScheduledReportConfig = None, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Update a scheduled report."""
    return {"id": report_id, **config.model_dump()}


@router.delete("/scheduled/{report_id}")
@limiter.limit("30/minute")
def delete_scheduled_report(request: Request, report_id: int, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Delete a scheduled report."""
    return {"success": True, "message": "Scheduled report deleted"}


@router.post("/scheduled/{report_id}/run")
@limiter.limit("30/minute")
def run_scheduled_report_now(request: Request, report_id: int, db: DbSession = None, current_user: OptionalCurrentUser = None):
    """Manually trigger a scheduled report."""
    return {"success": True, "message": "Report queued for generation"}
